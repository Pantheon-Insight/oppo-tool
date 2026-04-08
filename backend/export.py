from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from models import ExtractionResult
from datetime import datetime


def generate_xlsx(result: ExtractionResult) -> BytesIO:
    """
    Generate XLSX file from extraction results.
    Uses navy header, alternating row colors, and proper formatting.
    Returns BytesIO object ready to be streamed.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Attacks"

    # Define styles
    navy_fill = PatternFill(start_color="000033", end_color="000033", fill_type="solid")
    navy_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    white_font = Font(color="000000", size=11)
    gray_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC")
    )

    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # Create header row
    headers = [
        "No.",
        "Category",
        "Attack",
        "Key Detail",
        "Severity",
        "Best Universe",
        "Secondary Universe",
        "Tertiary Universe",
        "Notes"
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = navy_fill
        cell.font = navy_font
        cell.alignment = header_alignment
        cell.border = thin_border

    # Add data rows with alternating colors
    for row_num, attack in enumerate(result.attacks, 2):
        is_even_row = (row_num - 2) % 2 == 0
        row_fill = white_fill if is_even_row else gray_fill

        row_data = [
            attack.number,
            attack.category,
            attack.attack,
            attack.key_detail,
            attack.severity.value,
            attack.best_universe or "",
            attack.secondary_universe or "",
            attack.tertiary_universe or "",
            attack.notes or ""
        ]

        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.fill = row_fill
            cell.font = white_font
            cell.border = thin_border

            # Set alignment based on column
            if col_num == 1:  # Number
                cell.alignment = center_alignment
            elif col_num == 5:  # Severity
                cell.alignment = center_alignment
            else:
                cell.alignment = left_alignment

    # Set column widths
    column_widths = {
        "A": 6,   # No.
        "B": 18,  # Category
        "C": 28,  # Attack
        "D": 30,  # Key Detail
        "E": 12,  # Severity
        "F": 20,  # Best Universe
        "G": 20,  # Secondary Universe
        "H": 20,  # Tertiary Universe
        "I": 20   # Notes
    }

    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Set header row height
    ws.row_dimensions[1].height = 30

    # Add metadata sheet
    metadata_ws = wb.create_sheet("Metadata")
    metadata_ws["A1"] = "Subject"
    metadata_ws["B1"] = result.subject
    metadata_ws["A2"] = "Filename"
    metadata_ws["B2"] = result.filename
    metadata_ws["A3"] = "Total Attacks"
    metadata_ws["B3"] = result.total_attacks
    metadata_ws["A4"] = "Export Date"
    metadata_ws["B4"] = datetime.now().isoformat()
    metadata_ws["A5"] = "Job ID"
    metadata_ws["B5"] = result.job_id
    metadata_ws["A6"] = "Universes Used"
    metadata_ws["B6"] = ", ".join(result.universes_used) if result.universes_used else "N/A"

    # Format metadata sheet
    for cell in ["A1", "A2", "A3", "A4", "A5", "A6"]:
        metadata_ws[cell].fill = navy_fill
        metadata_ws[cell].font = navy_font

    metadata_ws.column_dimensions["A"].width = 20
    metadata_ws.column_dimensions["B"].width = 50

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output
